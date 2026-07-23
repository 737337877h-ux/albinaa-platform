# -*- coding: utf-8 -*-
"""
AlBinaa Credit & Collection Platform — Excel Statement Parser & Importer
=========================================================================
يقرأ ملف كشف الحساب التحليلي (بنية الكتل المتكررة) ويستورده بشكل Idempotent:
- الاستيراد الأول: إنشاء العملاء والحسابات والحركات.
- إعادة استيراد نفس الملف: صفر تكرار (عملاء أو حركات).
- دمج الكتل المجزأة لنفس (العميل، العملة) تلقائيًا.
- مطابقة الرصيد المحسوب مع الرصيد المعلن في الملف حيثما توفر.

ملاحظة تشغيلية: منطق الاستيراد مكتوب بصيغة SQL قياسية تعمل على PostgreSQL
(الإنتاج) وSQLite (بيئة الاختبار الحالية). القيود الفريدة متطابقة في الحالتين.

line_hash (النسخة المعدلة v2 — انظر تقرير المراجعة):
    SHA256(customer_code|currency|date|doc_type|doc_number|reference|debit|credit|norm_desc|occurrence_ordinal)
حيث occurrence_ordinal = ترتيب الصف بين الصفوف المتطابقة تمامًا لنفس الحساب
(بترتيب ظهورها في الملف). هذا يمنع تكرار الحركة عند إعادة استيراد نفس الكشف
التراكمي، دون إسقاط حركتين حقيقيتين متطابقتين.
"""
import hashlib
import re
import sqlite3
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, date

from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# ثوابت بنية الملف (كما اكتُشفت من التحليل الفعلي — لا افتراضات)
# ----------------------------------------------------------------------------
LBL_CUSTOMER = 'رقم العميل'
LBL_CURRENCY = 'العملة'
LBL_COL_DATE = 'التاريخ'
LBL_FOREIGN = 'المبلغ الأجنبي'
LBL_OPENING = 'الرصيد الإفتتاحي'
LBL_TOTALS = 'إجمالي العمليات'
LBL_BAL_CREDIT = 'إجمالي الرصيد لكم'    # رصيد دائن (نحن مدينون للعميل)
LBL_BAL_DEBIT = 'إجمالي الرصيد عليكم'   # رصيد مدين (العميل مدين لنا)
LBL_BAL_ZERO = 'الرصيد الحالي'          # يظهر فقط عندما الرصيد = 0

CURRENCY_MAP = {'YR': 'YER', 'SR': 'SAR', '$': 'USD'}


def normalize_text(s):
    """توحيد نص للمقارنة/البصمة: إزالة مسافات زائدة وتوحيد الأشكال."""
    if s is None:
        return ''
    s = str(s)
    s = unicodedata.normalize('NFKC', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def normalize_name_for_match(s):
    """تطبيع اسم العميل لأغراض كشف التشابه فقط (لا يُستخدم للدمج التلقائي)."""
    s = normalize_text(s)
    s = re.sub('[إأآا]', 'ا', s)
    s = re.sub('ى', 'ي', s)
    s = re.sub('ة', 'ه', s)
    s = re.sub(r'\s+', ' ', s)
    return s


# ----------------------------------------------------------------------------
# نموذج البيانات المستخرجة
# ----------------------------------------------------------------------------
@dataclass
class Txn:
    row_number: int          # رقم الصف في الملف الأصلي (1-based) للتدقيق
    tx_date: object
    doc_type: str
    doc_number: object
    description: str
    reference: object
    debit: float
    credit: float
    line_hash: str = ''


@dataclass
class Account:
    """حساب = (عميل، عملة) بعد دمج الكتل المجزأة."""
    customer_code: int
    customer_name: str
    currency_raw: str
    currency_name: str
    opening_debit: float = 0.0
    opening_credit: float = 0.0
    transactions: list = field(default_factory=list)
    declared_balance: float = None   # القيمة الموقعة: موجب=مدين علينا، سالب=دائن لنا
    declared_label: str = None
    fragments: int = 0
    parse_warnings: list = field(default_factory=list)

    @property
    def currency(self):
        return CURRENCY_MAP.get(self.currency_raw, self.currency_raw)

    @property
    def computed_balance(self):
        td = sum(t.debit for t in self.transactions)
        tc = sum(t.credit for t in self.transactions)
        return (self.opening_debit + td) - (self.opening_credit + tc)


@dataclass
class ParseResult:
    accounts: dict = field(default_factory=dict)   # (code, ccy) -> Account
    errors: list = field(default_factory=list)     # (row, message, raw)
    skipped_rows: list = field(default_factory=list)
    stats: dict = field(default_factory=dict)


# ----------------------------------------------------------------------------
# المرحلة 1: القارئ الهيكلي (Block Parser)
# ----------------------------------------------------------------------------
def parse_workbook(path, sheet_name=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    sn = sheet_name or wb.sheetnames[0]
    ws = wb[sn]
    rows = [(idx + 1, r) for idx, r in enumerate(ws.iter_rows(values_only=True))]
    n = len(rows)

    res = ParseResult()
    order = []
    i = 0
    while i < n:
        rownum, r = rows[i]
        if r[0] == LBL_CUSTOMER:
            cust_code, cust_name = r[1], normalize_text(r[3])
            i += 1
            # تجاوز صفوف "0" الثلاثة غامضة المعنى (قيمتها صفر دائمًا — موثقة بالتقرير)
            while i < n and rows[i][1][1] is not None and all(rows[i][1][k] is None for k in (0, 2, 3, 4, 5, 6)):
                if rows[i][1][1] != 0:
                    res.errors.append((rows[i][0], 'قيمة غير صفرية في صف القالب الغامض', rows[i][1]))
                i += 1
            # سطر العملة
            ccy_raw = ccy_name = None
            if i < n and rows[i][1][0] == LBL_CURRENCY:
                ccy_raw, ccy_name = rows[i][1][2], normalize_text(rows[i][1][3])
                i += 1
            else:
                res.errors.append((rows[i][0] if i < n else rownum, 'سطر العملة مفقود — الكتلة مستبعدة', rows[i][1] if i < n else None))
                continue
            # سطر "المبلغ الأجنبي" الاختياري ثم سطر عناوين الأعمدة
            if i < n and rows[i][1][5] == LBL_FOREIGN:
                i += 1
            if i < n and rows[i][1][0] == LBL_COL_DATE:
                i += 1
            else:
                res.errors.append((rows[i][0] if i < n else rownum, 'سطر عناوين الأعمدة مفقود', rows[i][1] if i < n else None))
            # الرصيد الافتتاحي (يظهر فقط في أول كتلة للحساب؛ الكتل المجزأة تكرره بنفس القيمة)
            od = oc = 0.0
            has_opening = False
            if i < n and rows[i][1][3] == LBL_OPENING:
                od = float(rows[i][1][5] or 0)
                oc = float(rows[i][1][6] or 0)
                has_opening = True
                i += 1

            key = (cust_code, ccy_raw)
            if key not in res.accounts:
                acc = Account(cust_code, cust_name, ccy_raw, ccy_name, od, oc)
                res.accounts[key] = acc
                order.append(key)
            else:
                acc = res.accounts[key]
                # كتلة مجزأة: نتأكد أن الافتتاحي المكرر مطابق للأول ولا نجمعه مرتين
                if has_opening and (abs(acc.opening_debit - od) > 0.005 or abs(acc.opening_credit - oc) > 0.005):
                    acc.parse_warnings.append(
                        f'رصيد افتتاحي مختلف في كتلة مجزأة (الصف {rownum}): '
                        f'({acc.opening_debit},{acc.opening_credit}) مقابل ({od},{oc})')
                if normalize_text(cust_name) != normalize_text(acc.customer_name):
                    acc.parse_warnings.append(f'اسم مختلف لنفس الكود في الصف {rownum}: "{cust_name}"')
            acc.fragments += 1

            # الحركات حتى سطر الإجمالي أو كتلة جديدة
            while i < n:
                rn2, rr = rows[i]
                if rr[3] == LBL_TOTALS:
                    i += 1
                    if i < n:
                        lbl = rows[i][1][2]
                        if lbl == LBL_BAL_DEBIT:
                            acc.declared_balance = float(rows[i][1][5] or 0)
                            acc.declared_label = lbl
                            i += 1
                        elif lbl == LBL_BAL_CREDIT:
                            acc.declared_balance = -float(rows[i][1][6] or 0)
                            acc.declared_label = lbl
                            i += 1
                        elif lbl == LBL_BAL_ZERO:
                            acc.declared_balance = 0.0
                            acc.declared_label = lbl
                            i += 1
                    break
                elif rr[0] == LBL_CUSTOMER:
                    break
                elif all(v is None for v in rr):
                    res.skipped_rows.append((rn2, 'صف فارغ'))
                    i += 1
                    continue
                else:
                    d = float(rr[5] or 0)
                    c = float(rr[6] or 0)
                    if d < 0 or c < 0:
                        res.errors.append((rn2, 'مبلغ سالب — الصف مستبعد', rr))
                        i += 1
                        continue
                    if d > 0 and c > 0:
                        res.errors.append((rn2, 'مدين ودائن معًا — الصف مستبعد للمراجعة', rr))
                        i += 1
                        continue
                    acc.transactions.append(Txn(
                        row_number=rn2,
                        tx_date=rr[0],
                        doc_type=normalize_text(rr[1]),
                        doc_number=rr[2],
                        description=normalize_text(rr[3]),
                        reference=rr[4],
                        debit=d, credit=c))
                    i += 1
        else:
            if any(v is not None for v in r):
                res.errors.append((rownum, 'صف خارج أي كتلة — مستبعد', r))
            else:
                res.skipped_rows.append((rownum, 'صف فارغ'))
            i += 1

    # حساب line_hash v2 مع الترتيب التسلسلي للحركات المتطابقة
    for key in order:
        acc = res.accounts[key]
        seen = Counter()
        for t in acc.transactions:
            natural = '|'.join([
                str(acc.customer_code), acc.currency,
                t.tx_date.date().isoformat() if isinstance(t.tx_date, datetime) else str(t.tx_date),
                t.doc_type, normalize_text(t.doc_number), normalize_text(t.reference),
                f'{t.debit:.4f}', f'{t.credit:.4f}', t.description,
            ])
            seen[natural] += 1
            t.line_hash = hashlib.sha256(f'{natural}|occ={seen[natural]}'.encode('utf-8')).hexdigest()

    res.stats = {
        'accounts': len(res.accounts),
        'customers': len({k[0] for k in res.accounts}),
        'transactions': sum(len(a.transactions) for a in res.accounts.values()),
        'fragmented_accounts': sum(1 for a in res.accounts.values() if a.fragments > 1),
        'errors': len(res.errors),
        'empty_rows_skipped': len(res.skipped_rows),
    }
    return res


# ----------------------------------------------------------------------------
# المرحلة 2: مخطط قاعدة الاختبار (مطابق منطقيًا لمخطط PostgreSQL الإنتاجي)
# ----------------------------------------------------------------------------
TEST_SCHEMA = """
CREATE TABLE IF NOT EXISTS currencies (
    code TEXT PRIMARY KEY, name_ar TEXT NOT NULL, symbol TEXT, active INTEGER DEFAULT 1);

CREATE TABLE IF NOT EXISTS document_types (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    effect TEXT NOT NULL CHECK(effect IN ('debit','credit','mixed')),
    active INTEGER DEFAULT 1);

CREATE TABLE IF NOT EXISTS customers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    external_customer_code TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    name_normalized TEXT NOT NULL,
    account_number TEXT,
    created_by_import_job INTEGER,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')));

CREATE TABLE IF NOT EXISTS customer_balances (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER NOT NULL REFERENCES customers(id),
    currency_code TEXT NOT NULL REFERENCES currencies(code),
    opening_debit REAL NOT NULL DEFAULT 0,
    opening_credit REAL NOT NULL DEFAULT 0,
    accounting_balance REAL NOT NULL DEFAULT 0,
    declared_balance REAL,
    declared_label TEXT,
    last_import_job_id INTEGER,
    updated_at TEXT DEFAULT (datetime('now')),
    UNIQUE(customer_id, currency_code));

CREATE TABLE IF NOT EXISTS import_jobs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    file_name TEXT NOT NULL,
    file_hash TEXT NOT NULL,
    imported_at TEXT DEFAULT (datetime('now')),
    rows_total INTEGER, txns_in_file INTEGER,
    customers_new INTEGER, customers_updated INTEGER,
    txns_inserted INTEGER, txns_skipped_duplicate INTEGER,
    errors_count INTEGER, status TEXT DEFAULT 'completed');

CREATE TABLE IF NOT EXISTS imported_transactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER NOT NULL REFERENCES customers(id),
    currency_code TEXT NOT NULL REFERENCES currencies(code),
    document_type_id INTEGER NOT NULL REFERENCES document_types(id),
    tx_date TEXT NOT NULL,
    document_number TEXT,
    description TEXT,
    reference_number TEXT,
    debit REAL NOT NULL DEFAULT 0,
    credit REAL NOT NULL DEFAULT 0,
    line_hash TEXT NOT NULL UNIQUE,
    source_row_number INTEGER,
    import_job_id INTEGER NOT NULL REFERENCES import_jobs(id));

CREATE TABLE IF NOT EXISTS balance_snapshots (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id INTEGER NOT NULL,
    currency_code TEXT NOT NULL,
    balance REAL NOT NULL,
    import_job_id INTEGER NOT NULL,
    snapshot_at TEXT DEFAULT (datetime('now')));

CREATE TABLE IF NOT EXISTS potential_duplicate_customers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_a_id INTEGER NOT NULL REFERENCES customers(id),
    customer_b_id INTEGER NOT NULL REFERENCES customers(id),
    match_reason TEXT NOT NULL,
    review_status TEXT NOT NULL DEFAULT 'pending',
    UNIQUE(customer_a_id, customer_b_id));
"""


def import_file(conn, path, parse_result=None):
    """استيراد Idempotent. يعيد قاموس إحصاءات."""
    res = parse_result or parse_workbook(path)
    cur = conn.cursor()

    with open(path, 'rb') as f:
        file_hash = hashlib.sha256(f.read()).hexdigest()

    already = cur.execute('SELECT id, imported_at FROM import_jobs WHERE file_hash=?',
                          (file_hash,)).fetchone()
    file_seen_before = already is not None

    cur.execute("""INSERT INTO import_jobs(file_name, file_hash, rows_total, txns_in_file, errors_count)
                   VALUES (?,?,?,?,?)""",
                (path.split('/')[-1], file_hash,
                 res.stats.get('empty_rows_skipped', 0), res.stats['transactions'],
                 res.stats['errors']))
    job_id = cur.lastrowid

    # العملات وأنواع المستندات (upsert)
    for raw, iso in CURRENCY_MAP.items():
        pass
    ccy_names = {}
    for acc in res.accounts.values():
        ccy_names[acc.currency] = acc.currency_name
    for code, name in ccy_names.items():
        cur.execute('INSERT OR IGNORE INTO currencies(code, name_ar) VALUES (?,?)', (code, name))

    DEBIT_TYPES = {'فاتورة المبيعات آجل', 'سند صرف نقدي', 'سند صرف بنكي', 'أمر توريد مخزني... (مؤقت)'}
    doc_type_ids = {}
    all_doc_types = set()
    for acc in res.accounts.values():
        for t in acc.transactions:
            all_doc_types.add(t.doc_type)
    for dt in sorted(all_doc_types):
        cur.execute('INSERT OR IGNORE INTO document_types(name, effect) VALUES (?,?)', (dt, 'mixed'))
    for dt_id, name in cur.execute('SELECT id, name FROM document_types'):
        doc_type_ids[name] = dt_id

    customers_new = customers_updated = 0
    txns_inserted = txns_skipped = 0

    for (code, ccy_raw), acc in res.accounts.items():
        code_s = str(code)
        row = cur.execute('SELECT id, name FROM customers WHERE external_customer_code=?',
                          (code_s,)).fetchone()
        if row is None:
            cur.execute("""INSERT INTO customers(external_customer_code, name, name_normalized, created_by_import_job)
                           VALUES (?,?,?,?)""",
                        (code_s, acc.customer_name, normalize_name_for_match(acc.customer_name), job_id))
            customer_id = cur.lastrowid
            customers_new += 1
        else:
            customer_id = row[0]
            if row[1] != acc.customer_name:
                cur.execute("UPDATE customers SET name=?, name_normalized=?, updated_at=datetime('now') WHERE id=?",
                            (acc.customer_name, normalize_name_for_match(acc.customer_name), customer_id))
            customers_updated += 1

        # الحركات — القيد الفريد على line_hash يمنع التكرار
        for t in acc.transactions:
            try:
                cur.execute("""INSERT INTO imported_transactions
                    (customer_id, currency_code, document_type_id, tx_date, document_number,
                     description, reference_number, debit, credit, line_hash, source_row_number, import_job_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (customer_id, acc.currency, doc_type_ids[t.doc_type],
                     t.tx_date.date().isoformat() if isinstance(t.tx_date, datetime) else str(t.tx_date),
                     normalize_text(t.doc_number), t.description, normalize_text(t.reference),
                     t.debit, t.credit, t.line_hash, t.row_number, job_id))
                txns_inserted += 1
            except sqlite3.IntegrityError:
                txns_skipped += 1

        # الرصيد المحاسبي (upsert على القيد الفريد customer_id+currency)
        bal = acc.computed_balance
        cur.execute("""INSERT INTO customer_balances
                (customer_id, currency_code, opening_debit, opening_credit,
                 accounting_balance, declared_balance, declared_label, last_import_job_id)
            VALUES (?,?,?,?,?,?,?,?)
            ON CONFLICT(customer_id, currency_code) DO UPDATE SET
                opening_debit=excluded.opening_debit,
                opening_credit=excluded.opening_credit,
                accounting_balance=excluded.accounting_balance,
                declared_balance=excluded.declared_balance,
                declared_label=excluded.declared_label,
                last_import_job_id=excluded.last_import_job_id,
                updated_at=datetime('now')""",
            (customer_id, acc.currency, acc.opening_debit, acc.opening_credit,
             bal, acc.declared_balance, acc.declared_label, job_id))
        cur.execute("""INSERT INTO balance_snapshots(customer_id, currency_code, balance, import_job_id)
                       VALUES (?,?,?,?)""", (customer_id, acc.currency, bal, job_id))

    # كشف تشابه الأسماء (تنبيه فقط — لا دمج أبدًا)
    dup_pairs = 0
    rows = cur.execute('SELECT id, name_normalized FROM customers ORDER BY id').fetchall()
    by_norm = defaultdict(list)
    for cid, nn in rows:
        by_norm[nn].append(cid)
    for nn, ids in by_norm.items():
        if len(ids) > 1:
            for a in range(len(ids)):
                for b in range(a + 1, len(ids)):
                    cur.execute("""INSERT OR IGNORE INTO potential_duplicate_customers
                                   (customer_a_id, customer_b_id, match_reason)
                                   VALUES (?,?,?)""",
                                (ids[a], ids[b], 'تطابق اسم تام بعد التطبيع مع اختلاف الكود'))
                    dup_pairs += 1

    cur.execute("""UPDATE import_jobs SET customers_new=?, customers_updated=?,
                   txns_inserted=?, txns_skipped_duplicate=? WHERE id=?""",
                (customers_new, customers_updated, txns_inserted, txns_skipped, job_id))
    conn.commit()
    return {
        'job_id': job_id, 'file_hash': file_hash, 'file_seen_before': file_seen_before,
        'customers_new': customers_new, 'customers_updated': customers_updated,
        'txns_inserted': txns_inserted, 'txns_skipped_duplicate': txns_skipped,
        'accounts': res.stats['accounts'], 'parse_errors': res.stats['errors'],
        'fragmented_accounts_merged': res.stats['fragmented_accounts'],
        'duplicate_name_pairs_flagged': dup_pairs,
    }


def verify_balances(res):
    """مطابقة الرصيد المحسوب مع الرصيد المعلن حيثما وُجد سطر إجمالي في الملف."""
    checked = matched = mismatched = no_declared = 0
    mismatches = []
    for acc in res.accounts.values():
        if acc.declared_balance is None:
            no_declared += 1
            continue
        checked += 1
        if abs(acc.computed_balance - acc.declared_balance) <= 0.5:
            matched += 1
        else:
            mismatched += 1
            mismatches.append((acc.customer_code, acc.currency,
                               acc.computed_balance, acc.declared_balance))
    return {'checked': checked, 'matched': matched, 'mismatched': mismatched,
            'no_declared_line': no_declared, 'mismatch_details': mismatches}


if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/عملاء_تحليلي_16-07-2026.xlsx'
    db = sys.argv[2] if len(sys.argv) > 2 else '/home/claude/albinaa_test.db'
    conn = sqlite3.connect(db)
    conn.executescript(TEST_SCHEMA)
    res = parse_workbook(src)
    print('PARSE:', res.stats)
    print('IMPORT:', import_file(conn, src, res))
    print('VERIFY:', verify_balances(res))
